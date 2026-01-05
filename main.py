# main.py

import re
import os
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ================== TOKEN (تعديل هنا) ==================
# هنا نقوم بسحب التوكن من إعدادات Railway التي قمت بضبطها مسبقاً
TOKEN = os.getenv("8496832736:AAGMC-tAQTV6U-VZh7ec6dOcfykfhE2E6pE")

# ================== STYLES =================
blue_fill = PatternFill("solid", fgColor="ADD8E6")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
light_blue_fill = PatternFill("solid", fgColor="00B0F0")
dark_blue_fill = PatternFill("solid", fgColor="0070C0")

header_font = Font(bold=True)
title_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ================== HELPERS =================
def parse_loops_diameter(text):
    text = text.replace(" ", "")
    if "*" not in text:
        return None, None, None
    try:
        a, b = map(float, text.split("*"))
        loops, diameter = (a, b) if a <= b else (b, a)
        result = round(loops * diameter * 3.14 / 100, 3)
        return loops, diameter, result
    except:
        return None, None, None


def extract_cable_type(line):
    m = re.search(r"\d+\s*[Ff]\s*\d+", line)
    return m.group(0) if m else None


def new_cable():
    return {"type": "", "loops": "", "diameter": "", "result": ""}


# ================== PARSER =================
def parse_data_smart(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        return None

    data = {
        "FDH": lines[0],
        "poles": [],
        "handhole": [],
        "fdh_loop": [],
    }

    section = None
    current_entry = None
    current_cable = None

    for ln in lines[1:]:
        m_p = re.match(r"^[Pp](\d+)", ln)
        if m_p:
            num = int(m_p.group(1))
            section = "poles"
            current_entry = {
                "id": f"P{num}",
                "sort": num,
                "cables": [],
            }
            data["poles"].append(current_entry)
            current_cable = None
            continue

        m_h = re.match(r"^[Hh](\d+)", ln)
        if m_h:
            num = int(m_h.group(1))
            section = "handhole"
            current_entry = {
                "id": f"H{num}",
                "sort": num,
                "cables": [],
            }
            data["handhole"].append(current_entry)
            current_cable = None
            continue

        m_f = re.match(r"^[Ff][Dd][Hh](\d*)", ln)
        if m_f:
            num = int(m_f.group(1)) if m_f.group(1) else 0
            section = "fdh_loop"
            current_entry = {
                "id": f"FDH{num}" if num else "FDH",
                "sort": num,
                "cables": [],
            }
            data["fdh_loop"].append(current_entry)
            current_cable = None
            continue

        if not section or not current_entry:
            continue

        cable_type = extract_cable_type(ln)
        if cable_type:
            current_cable = new_cable()
            current_cable["type"] = cable_type
            current_entry["cables"].append(current_cable)
            continue

        if "*" in ln:
            if not current_cable:
                current_cable = new_cable()
                current_entry["cables"].append(current_cable)
            loops, dia, res = parse_loops_diameter(ln)
            current_cable.update(
                {"loops": loops, "diameter": dia, "result": res}
            )
            continue

    data["poles"].sort(key=lambda x: x["sort"])
    data["handhole"].sort(key=lambda x: x["sort"])
    data["fdh_loop"].sort(key=lambda x: x["sort"])

    return data
# ================== EXCEL =================
def make_sheet(ws, title, entries, id_name):
    ws["A1"] = title
    ws.merge_cells("A1:E1")
    ws["A1"].fill = blue_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    headers = [
        id_name,
        "Cable Type",
        "Diameter (cm)",
        "Loops",
        "Length (m)",
    ]
    fills = [yellow_fill, yellow_fill, light_blue_fill, light_blue_fill, dark_blue_fill]

    for i, (h, f) in enumerate(zip(headers, fills), start=1):
        c = ws.cell(2, i, h)
        c.fill = f
        c.font = header_font
        c.alignment = center_align
        c.border = border

    r = 3
    for e in entries:
        first = True
        for cbl in e["cables"]:
            ws.cell(r, 1, e["id"] if first else "").border = border
            ws.cell(r, 2, cbl["type"]).border = border
            ws.cell(r, 3, cbl["diameter"]).border = border
            ws.cell(r, 4, cbl["loops"]).border = border
            ws.cell(r, 5, cbl["result"]).border = border
            for c in range(1, 6):
                ws.cell(r, c).alignment = center_align
            r += 1
            first = False

    for col in "ABCDE":
        ws.column_dimensions[col].width = 24


def create_excel(data):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Poles"
    make_sheet(ws1, data["FDH"], data["poles"], "POLE ID")

    ws2 = wb.create_sheet("Handholes")
    make_sheet(ws2, data["FDH"], data["handhole"], "H.H ID")

    ws3 = wb.create_sheet("FDH Loops")
    make_sheet(ws3, data["FDH"], data["fdh_loop"], "FDH ID")

    name = re.sub(r"[^A-Za-z0-9_-]", "_", data["FDH"])
    file = f"{name}_loop_length.xlsx"
    wb.save(file)
    return file


# ================== BOT =================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 أرسل البيانات بأي ترتيب\n"
        "P6 / P10 / P7 …\n"
        "سأرتّبها تلقائياً مع بياناتها كاملة"
    )


async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        data = parse_data_smart(update.message.text)
        if not data:
            await update.message.reply_text("❌ بيانات غير مفهومة")
            return

        file = create_excel(data)
        with open(file, "rb") as f:
            await update.message.reply_document(f)
        os.remove(file)

    except Exception as e:
        await update.message.reply_text(f"❌ خطأ: {e}")


def main():
    # التأكد من أن التوكن موجود
    if not TOKEN:
        print("❌ Error: BOT_TOKEN variable not found in environment!")
        return

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))
    print("🤖 BOT IS RUNNING...")
    app.run_polling()


if __name__ == "main":
    main()
